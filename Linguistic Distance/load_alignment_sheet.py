from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from collections import defaultdict


class Word:
    def __init__(self, color, points, name):
        self.color = color
        self.points = points
        self.name = name
        color_dict[color] = self

#Initialize word types with default color-coding scheme
color_dict = {}
blank = '00000000' #hexadecimal color code for white/blank cell
cognate = Word(blank, 0, 'cognate')                         #blank by default
non_cognate = Word('FFFFFF00', 1, 'non-cognate')            #yellow by default
partial_cognate = Word('FFFFC000', 0.5, 'partial cognate')  #orange by default
false_friend = Word('FF92D050', 1, 'false friend')          #green by default
color_dict['0'] = cognate


def get_color(sheet, cell):
    """Returns the hexadecimal color value of a particular cell, e.g. cell 'K44'"""
    return str(sheet[cell].fill.start_color.index) #hexadecimal

def load_alignfile(excel_alignment):
    """Loads an  Excel file containing sentence/phrase alignments"""
    return load_workbook(excel_alignment, data_only = True)


class AlignFile:

    def __init__(self, filepath, sheet='Sheet',
             lang1_column='B', lang2_column='B',
             start_column='E', start_row=2):
        
        #Load the Excel file
        self.filepath = filepath
        self.wb = load_workbook(self.filepath)
        self.sh = self.wb[sheet]
        
        #Get lists of columns and rows in the spreadsheet
        self.columns = [i[0].column_letter for i in list(self.sh.columns)]
        self.start_column = start_column
        self.start_column_i = self.columns.index(self.start_column)
        self.align_columns = self.columns[self.start_column_i:]
        self.lang1_column = lang1_column
        self.lang2_column = lang2_column
        self.rows = list(range(1,len(list(self.sh.rows))+1))
        self.start_row = start_row
        
        #Get labels of languages
        self.lang1 = self.sh[f'{self.columns[self.start_column_i-1]}{self.start_row+2}'].value
        self.lang2 = self.sh[f'{self.columns[self.start_column_i-1]}{self.start_row+3}'].value

        #Generate dictionary with list of cell values in each line
        self.lines = defaultdict(lambda:[])
        index = 0
        for line in list(self.sh.values):
            index += 1
            line = list(line)
            length = len(line)
            new_line = []
            i = length - 1
            while i != 0:
                if line[i] != None:
                    break
                else:
                    i -= 1
            new_line = line[self.start_column_i:i+1] #exclude empty cells at the end of the line
            self.lines[index].append(new_line)
    
        #Generate dictionary with a list of all cells filled with each color
        self.colors = defaultdict(lambda:[])
        for column in self.columns:
            for row in self.rows:
                cell = f'{column}{row}'
                color = get_color(self.sh, cell)
                self.colors[color].append(cell)
        self.align_color = get_color(self.sh, f'{self.start_column}{self.start_row}')
    
        #Check that all of the colored cells use one of the recognized colors
        for color in self.colors:
            if color not in color_dict:
                if color != self.align_color:
                    print(f'Error: word type coding for color {color} (e.g. cell {self.colors[color][0]}) is not defined)')
    
        #Get list of start lines of alignments
        self.start_lines = [line for line in self.lines 
                            if len(self.lines[line][0]) > 0 #not empty line 
                            if self.lines[line][0][0] == 1  #first element is 1, i.e. index 
                            if get_color(self.sh, f'{self.start_column}{line}') == self.align_color]
        
        #Get dictionary of phrase pairs indexed by each starting line
        self.phrases = {start_line:(self.sh[f'{lang1_column}{start_line+2}'].value, 
                                    self.sh[f'{lang2_column}{start_line+3}'].value) 
                        for start_line in self.start_lines}
        
        
    def separate_coordinates(self, cell):
        """Separates the row and column cell coordinates
        Returns a tuple (column, row)"""
        j = 0
        while cell[:j+1] in self.columns:
            j += 1
        c = cell[:j]
        r = int(cell[j:])
        return c, r
        
    
    def change_cell_values(self, cell_value_dict):
        """cell_value_dict : dictionary with cell coordinates as keys and new cell values as values
        Changes the cells of self.sh according to cell_value_dict"""
        
        for cell in cell_value_dict:
            self.sh[cell] = cell_value_dict[cell]
    
    
    def align_length(self, start_line):
        """Calculates the total alignment length as well as the number of 
        L1 and L2 words/units.
        Returns result as a tuple (total_length, l1_length, l2_length)"""
        l1_line = start_line + 2
        l2_line = start_line + 3
        measure_line = start_line + 4
        
                    
        total_length, l1_length, l2_length = 0, 0, 0
        end_align = False
        while end_align == False:
            column = self.align_columns[total_length]            
            l1_word = self.sh[f'{column}{l1_line}'].value
            l2_word = self.sh[f'{column}{l2_line}'].value
            
            if ((l1_word == None) and (l2_word == None)):
                end_align = True
            else:
                total_length += 1                
                if l1_word != None:
                    l1_length += 1
                if l2_word != None:
                    l2_length += 1
                    
        return total_length, l1_length, l2_length
    
    
    def null_alignments(self, start_lines=None):
        """Returns a list of tuples (cell, L1Word, L2Word) of empty alignments"""
        
        if start_lines == None:
            start_lines = self.start_lines
        
        null_alignments = []
        for start_line in start_lines:
            l1_line = start_line + 2
            l2_line = start_line + 3
            measure_line = start_line + 4
                        
            length = self.align_length(start_line)[0]
            for i in range(length):
                column = self.align_columns[i]
                l1_word = self.sh[f'{column}{l1_line}'].value
                l2_word = self.sh[f'{column}{l2_line}'].value
                measure = self.sh[f'{column}{measure_line}'].value
        
                if measure == None:
                    cell = f'{column}{measure_line}'
                    null_alignments.append((cell, l1_word, l2_word))
                    
        return null_alignments
    
    
    def change_align_color(self, new_color, start_lines=None):
        """Changes the color of the cells surrounding an alignment, given the 
        alignment's starting line. If no starting lines are specified,
        all start_lines in the alignment sheet are used by default."""
        
        if start_lines == None:
            start_lines = self.start_lines
        
        new_color = PatternFill(start_color=new_color, 
                                end_color=new_color, 
                                fill_type='solid')
            
        for start_line in start_lines:
            align_length = self.align_length(start_line)[0]
            for row in [start_line, start_line+1, start_line+4]:
                for i in range(align_length+3):
                    column = self.columns[(self.start_column_i-1)+i]
                    cell = self.sh[f'{column}{row}']
                    cell.fill = new_color
                    
        
    